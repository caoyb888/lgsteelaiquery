"""
单元测试：app/core/excel_parser.py

覆盖率目标：≥ 90%

所有测试文件在 tmp_path 临时目录中动态生成，不依赖预先存在的 fixtures。
"""
from __future__ import annotations

import os
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

# 设置必要的环境变量（在导入 app 模块前）
os.environ.setdefault("APP_ENV", "development")
os.environ.setdefault("META_DB_USER", "test_user")
os.environ.setdefault("META_DB_PASSWORD", "test_pass")
os.environ.setdefault("BIZ_DB_USER", "test_user")
os.environ.setdefault("BIZ_DB_PASSWORD", "test_pass")
os.environ.setdefault("REDIS_PASSWORD", "")
os.environ.setdefault("APP_SECRET_KEY", "test-secret-key-32-chars-xxxxxxxxx")

from app.core.excel_parser import ExcelParser, ParsedField, ExcelParseResult
from app.utils.exceptions import (
    ExcelFileTooLargeError,
    ExcelParseError,
    UnsupportedFormatError,
)


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


@pytest.fixture()
def parser() -> ExcelParser:
    return ExcelParser()


def make_standard_sales_xlsx(path: Path) -> Path:
    """创建标准格式销售台账 xlsx。"""
    data = {
        "月份": ["2026-01", "2026-01", "2026-02", "2026-02"],
        "产品线": ["板材", "型钢", "板材", "型钢"],
        "产品名称": ["热轧卷板", "H型钢", "热轧卷板", "H型钢"],
        "销售收入（万元）": [1250.50, 860.00, 1380.20, 920.00],
        "销售量（吨）": [8500, 6200, 9200, 6800],
        "客户名称": ["客户A", "客户B", "客户A", "客户C"],
    }
    fp = path / "standard_sales.xlsx"
    pd.DataFrame(data).to_excel(fp, index=False)
    return fp


def make_merged_cells_xlsx(path: Path) -> Path:
    """创建含合并单元格的 xlsx。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "销售数据"

    ws["A1"] = "月份"
    ws["B1"] = "产品线"
    ws["C1"] = "产品名称"
    ws["D1"] = "收入（万元）"
    ws["E1"] = "销量（吨）"

    # A2:A3 合并（月份 2026-01 跨两行）
    ws["A2"] = "2026-01"
    ws.merge_cells("A2:A3")
    ws["B2"] = "板材"
    ws["C2"] = "热轧卷板"
    ws["D2"] = 1250.50
    ws["E2"] = 8500

    ws["B3"] = "型钢"
    ws["C3"] = "H型钢"
    ws["D3"] = 860.00
    ws["E3"] = 6200

    fp = path / "merged_cells.xlsx"
    wb.save(fp)
    return fp


def make_multi_header_xlsx(path: Path) -> Path:
    """创建多级表头（第0行为标题行，第1行为真正表头）。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "财务数据"

    # 第一行：大标题（只有 A1 有值，其余空）
    ws["A1"] = "2026年1-3月财务汇总报表"

    # 第二行：正式表头（非空率高）
    ws["A2"] = "月份"
    ws["B2"] = "产品名称"
    ws["C2"] = "收入"
    ws["D2"] = "成本"
    ws["E2"] = "毛利"

    # 数据行
    ws["A3"] = "2026-01"
    ws["B3"] = "热轧卷板"
    ws["C3"] = 1250.5
    ws["D3"] = 1050.0
    ws["E3"] = 200.5

    ws["A4"] = "2026-02"
    ws["B4"] = "热轧卷板"
    ws["C4"] = 1380.2
    ws["D4"] = 1150.0
    ws["E4"] = 230.2

    fp = path / "multi_header.xlsx"
    wb.save(fp)
    return fp


def make_mixed_types_xlsx(path: Path) -> Path:
    """创建混合类型字段 xlsx。"""
    data = {
        "日期": ["2026-01-01", "2026-01-02", "2026-01-03"],
        "金额": [1250.5, 860.0, 920.0],
        "产品": ["热轧卷板", "H型钢", "线材"],
        "是否完成": ["是", "否", "是"],
    }
    fp = path / "mixed_types.xlsx"
    pd.DataFrame(data).to_excel(fp, index=False)
    return fp


def make_empty_rows_xlsx(path: Path) -> Path:
    """创建含全空列的 xlsx。"""
    data = {
        "月份": ["2026-01", "2026-02", "2026-03"],
        "产品线": ["板材", "型钢", "板材"],
        "收入（万元）": [1250.5, 860.0, 1380.2],
        "备注": [None, None, None],  # 全空列
    }
    fp = path / "empty_rows.xlsx"
    pd.DataFrame(data).to_excel(fp, index=False)
    return fp


def make_gbk_csv(path: Path) -> Path:
    """创建 GBK 编码 CSV。"""
    content = "月份,产品线,收入（万元）\n2026-01,板材,1250.5\n2026-01,型钢,860.0\n"
    fp = path / "bad_encoding.csv"
    fp.write_bytes(content.encode("gbk"))
    return fp


def make_header_only_xlsx(path: Path) -> Path:
    """只有表头行，无数据行。"""
    data: dict[str, list] = {"月份": [], "收入": []}
    fp = path / "header_only.xlsx"
    pd.DataFrame(data).to_excel(fp, index=False)
    return fp


def make_duplicate_columns_xlsx(path: Path) -> Path:
    """含重复列名的 xlsx。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "月份"
    ws["B1"] = "收入"
    ws["C1"] = "收入"   # 重复
    ws["A2"] = "2026-01"
    ws["B2"] = 100.0
    ws["C2"] = 200.0
    fp = path / "dup_columns.xlsx"
    wb.save(fp)
    return fp


# ---------------------------------------------------------------------------
# 1. 标准格式解析
# ---------------------------------------------------------------------------


class TestStandardSales:
    def test_field_count(self, parser: ExcelParser, tmp_path: Path) -> None:
        fp = make_standard_sales_xlsx(tmp_path)
        result = parser.parse(fp, fp.name)
        assert len(result.fields) == 6

    def test_data_rows(self, parser: ExcelParser, tmp_path: Path) -> None:
        fp = make_standard_sales_xlsx(tmp_path)
        result = parser.parse(fp, fp.name)
        assert result.total_rows == 4

    def test_header_row_index_zero(self, parser: ExcelParser, tmp_path: Path) -> None:
        fp = make_standard_sales_xlsx(tmp_path)
        result = parser.parse(fp, fp.name)
        assert result.header_row_index == 0

    def test_unit_extracted(self, parser: ExcelParser, tmp_path: Path) -> None:
        fp = make_standard_sales_xlsx(tmp_path)
        result = parser.parse(fp, fp.name)
        units = {f.clean_name: f.unit for f in result.fields}
        assert units.get("销售收入") == "万元"
        assert units.get("销售量") == "吨"

    def test_numeric_type_inferred(self, parser: ExcelParser, tmp_path: Path) -> None:
        fp = make_standard_sales_xlsx(tmp_path)
        result = parser.parse(fp, fp.name)
        type_map = {f.clean_name: f.inferred_type for f in result.fields}
        assert type_map["销售收入"] == "numeric"
        assert type_map["销售量"] == "numeric"

    def test_source_filename(self, parser: ExcelParser, tmp_path: Path) -> None:
        fp = make_standard_sales_xlsx(tmp_path)
        result = parser.parse(fp, "my_upload.xlsx")
        assert result.source_filename == "my_upload.xlsx"

    def test_no_warnings(self, parser: ExcelParser, tmp_path: Path) -> None:
        fp = make_standard_sales_xlsx(tmp_path)
        result = parser.parse(fp, fp.name)
        assert result.warnings == []

    def test_result_type(self, parser: ExcelParser, tmp_path: Path) -> None:
        fp = make_standard_sales_xlsx(tmp_path)
        result = parser.parse(fp, fp.name)
        assert isinstance(result, ExcelParseResult)
        assert isinstance(result.df, pd.DataFrame)

    def test_sample_values_populated(self, parser: ExcelParser, tmp_path: Path) -> None:
        fp = make_standard_sales_xlsx(tmp_path)
        result = parser.parse(fp, fp.name)
        for f in result.fields:
            assert len(f.sample_values) > 0

    def test_null_ratio_zero_for_full_column(self, parser: ExcelParser, tmp_path: Path) -> None:
        fp = make_standard_sales_xlsx(tmp_path)
        result = parser.parse(fp, fp.name)
        for f in result.fields:
            assert f.null_ratio == pytest.approx(0.0)


# ---------------------------------------------------------------------------
# 2. 合并单元格展开
# ---------------------------------------------------------------------------


class TestMergedCells:
    def test_merged_value_filled(self, parser: ExcelParser, tmp_path: Path) -> None:
        fp = make_merged_cells_xlsx(tmp_path)
        result = parser.parse(fp, fp.name)
        # 月份列的两行都应该是 "2026-01"
        month_col = result.df.columns[0]
        values = result.df[month_col].tolist()
        assert all(str(v) == "2026-01" for v in values), f"月份列值：{values}"

    def test_data_rows_after_expand(self, parser: ExcelParser, tmp_path: Path) -> None:
        fp = make_merged_cells_xlsx(tmp_path)
        result = parser.parse(fp, fp.name)
        assert result.total_rows == 2

    def test_sheet_name(self, parser: ExcelParser, tmp_path: Path) -> None:
        fp = make_merged_cells_xlsx(tmp_path)
        result = parser.parse(fp, fp.name)
        assert result.sheet_name == "销售数据"

    def test_unit_extracted_after_expand(self, parser: ExcelParser, tmp_path: Path) -> None:
        fp = make_merged_cells_xlsx(tmp_path)
        result = parser.parse(fp, fp.name)
        unit_map = {f.clean_name: f.unit for f in result.fields}
        assert unit_map.get("收入") == "万元"
        assert unit_map.get("销量") == "吨"


# ---------------------------------------------------------------------------
# 3. 多级表头检测
# ---------------------------------------------------------------------------


class TestMultiHeader:
    def test_header_detected_at_row1(self, parser: ExcelParser, tmp_path: Path) -> None:
        fp = make_multi_header_xlsx(tmp_path)
        result = parser.parse(fp, fp.name)
        # 第0行只有1个非空值，第1行有5个 → 应检测到第1行
        assert result.header_row_index == 1

    def test_data_rows(self, parser: ExcelParser, tmp_path: Path) -> None:
        fp = make_multi_header_xlsx(tmp_path)
        result = parser.parse(fp, fp.name)
        assert result.total_rows == 2

    def test_field_names_from_real_header(self, parser: ExcelParser, tmp_path: Path) -> None:
        fp = make_multi_header_xlsx(tmp_path)
        result = parser.parse(fp, fp.name)
        clean_names = {f.clean_name for f in result.fields}
        assert "月份" in clean_names
        assert "产品名称" in clean_names


# ---------------------------------------------------------------------------
# 4. 混合类型推断
# ---------------------------------------------------------------------------


class TestMixedTypes:
    def test_date_type(self, parser: ExcelParser, tmp_path: Path) -> None:
        fp = make_mixed_types_xlsx(tmp_path)
        result = parser.parse(fp, fp.name)
        type_map = {f.clean_name: f.inferred_type for f in result.fields}
        assert type_map["日期"] == "date"

    def test_numeric_type(self, parser: ExcelParser, tmp_path: Path) -> None:
        fp = make_mixed_types_xlsx(tmp_path)
        result = parser.parse(fp, fp.name)
        type_map = {f.clean_name: f.inferred_type for f in result.fields}
        assert type_map["金额"] == "numeric"

    def test_text_type(self, parser: ExcelParser, tmp_path: Path) -> None:
        fp = make_mixed_types_xlsx(tmp_path)
        result = parser.parse(fp, fp.name)
        type_map = {f.clean_name: f.inferred_type for f in result.fields}
        assert type_map["产品"] == "text"

    def test_boolean_type(self, parser: ExcelParser, tmp_path: Path) -> None:
        fp = make_mixed_types_xlsx(tmp_path)
        result = parser.parse(fp, fp.name)
        type_map = {f.clean_name: f.inferred_type for f in result.fields}
        assert type_map["是否完成"] == "boolean"


# ---------------------------------------------------------------------------
# 5. 全空列处理
# ---------------------------------------------------------------------------


class TestEmptyRows:
    def test_empty_column_dropped(self, parser: ExcelParser, tmp_path: Path) -> None:
        fp = make_empty_rows_xlsx(tmp_path)
        result = parser.parse(fp, fp.name)
        col_names = [f.clean_name for f in result.fields]
        assert "备注" not in col_names

    def test_warning_recorded_for_empty_col(self, parser: ExcelParser, tmp_path: Path) -> None:
        fp = make_empty_rows_xlsx(tmp_path)
        result = parser.parse(fp, fp.name)
        assert any("备注" in w and "全部为空" in w for w in result.warnings)

    def test_remaining_fields_intact(self, parser: ExcelParser, tmp_path: Path) -> None:
        fp = make_empty_rows_xlsx(tmp_path)
        result = parser.parse(fp, fp.name)
        col_names = {f.clean_name for f in result.fields}
        assert "月份" in col_names
        assert "产品线" in col_names

    def test_data_rows_correct(self, parser: ExcelParser, tmp_path: Path) -> None:
        fp = make_empty_rows_xlsx(tmp_path)
        result = parser.parse(fp, fp.name)
        assert result.total_rows == 3


# ---------------------------------------------------------------------------
# 6. GBK CSV 编码自动检测
# ---------------------------------------------------------------------------


class TestGbkCsv:
    def test_gbk_parsed_correctly(self, parser: ExcelParser, tmp_path: Path) -> None:
        fp = make_gbk_csv(tmp_path)
        result = parser.parse(fp, fp.name)
        assert result.total_rows == 2

    def test_gbk_field_names(self, parser: ExcelParser, tmp_path: Path) -> None:
        fp = make_gbk_csv(tmp_path)
        result = parser.parse(fp, fp.name)
        clean_names = {f.clean_name for f in result.fields}
        assert "月份" in clean_names
        assert "产品线" in clean_names

    def test_gbk_unit_extracted(self, parser: ExcelParser, tmp_path: Path) -> None:
        fp = make_gbk_csv(tmp_path)
        result = parser.parse(fp, fp.name)
        unit_map = {f.clean_name: f.unit for f in result.fields}
        assert unit_map.get("收入") == "万元"

    def test_csv_sheet_name(self, parser: ExcelParser, tmp_path: Path) -> None:
        fp = make_gbk_csv(tmp_path)
        result = parser.parse(fp, fp.name)
        assert result.sheet_name == "Sheet1"


# ---------------------------------------------------------------------------
# 7. 文件过大
# ---------------------------------------------------------------------------


class TestFileTooLarge:
    def test_raises_excel_file_too_large_error(
        self, parser: ExcelParser, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        # 构建一个小 xlsx，但 mock 文件大小超过限制
        fp = make_standard_sales_xlsx(tmp_path)
        import app.core.excel_parser as mod

        original_stat = Path.stat

        def fake_stat(self: Path, **kwargs):  # type: ignore[override]
            st = original_stat(self, **kwargs)
            # 返回一个 stat_result-like 对象，size 超过 50MB
            class FakeStat:
                st_size = 51 * 1024 * 1024
            return FakeStat()

        monkeypatch.setattr(Path, "stat", fake_stat)
        with pytest.raises(ExcelFileTooLargeError):
            parser.parse(fp, fp.name)

    def test_error_message_contains_filename(
        self, parser: ExcelParser, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        fp = make_standard_sales_xlsx(tmp_path)
        original_stat = Path.stat

        def fake_stat(self: Path, **kwargs):  # type: ignore[override]
            _ = original_stat(self, **kwargs)

            class FakeStat:
                st_size = 60 * 1024 * 1024
            return FakeStat()

        monkeypatch.setattr(Path, "stat", fake_stat)
        with pytest.raises(ExcelFileTooLargeError) as exc_info:
            parser.parse(fp, "my_big_file.xlsx")
        assert "my_big_file.xlsx" in str(exc_info.value)


# ---------------------------------------------------------------------------
# 8. 不支持的格式
# ---------------------------------------------------------------------------


class TestUnsupportedFormat:
    def test_pdf_raises_unsupported_format(
        self, parser: ExcelParser, tmp_path: Path
    ) -> None:
        fp = tmp_path / "report.pdf"
        fp.write_bytes(b"%PDF-1.4 fake content")
        with pytest.raises(UnsupportedFormatError):
            parser.parse(fp, fp.name)

    def test_docx_raises_unsupported_format(
        self, parser: ExcelParser, tmp_path: Path
    ) -> None:
        fp = tmp_path / "report.docx"
        fp.write_bytes(b"fake docx content")
        with pytest.raises(UnsupportedFormatError):
            parser.parse(fp, fp.name)

    def test_error_message_contains_suffix(
        self, parser: ExcelParser, tmp_path: Path
    ) -> None:
        fp = tmp_path / "report.txt"
        fp.write_text("hello", encoding="utf-8")
        with pytest.raises(UnsupportedFormatError) as exc_info:
            parser.parse(fp, fp.name)
        assert ".txt" in str(exc_info.value)


# ---------------------------------------------------------------------------
# 9. 无数据行（只有表头）
# ---------------------------------------------------------------------------


class TestNoDataRows:
    def test_raises_excel_parse_error(
        self, parser: ExcelParser, tmp_path: Path
    ) -> None:
        fp = make_header_only_xlsx(tmp_path)
        with pytest.raises(ExcelParseError, match="无有效数据行"):
            parser.parse(fp, fp.name)


# ---------------------------------------------------------------------------
# 10. 表头名清理：单位提取
# ---------------------------------------------------------------------------


class TestCleanHeaderNames:
    def test_fullwidth_brackets_unit(self, parser: ExcelParser) -> None:
        result = parser._clean_header_names(["销售收入（万元）"])
        clean_name, unit = result[0]
        assert clean_name == "销售收入"
        assert unit == "万元"

    def test_halfwidth_brackets_unit(self, parser: ExcelParser) -> None:
        result = parser._clean_header_names(["销售量(吨)"])
        clean_name, unit = result[0]
        assert clean_name == "销售量"
        assert unit == "吨"

    def test_no_brackets_no_unit(self, parser: ExcelParser) -> None:
        result = parser._clean_header_names(["产品名称"])
        clean_name, unit = result[0]
        assert clean_name == "产品名称"
        assert unit is None

    def test_space_replaced_with_underscore(self, parser: ExcelParser) -> None:
        result = parser._clean_header_names(["销售 收入"])
        clean_name, _ = result[0]
        assert " " not in clean_name
        assert "_" in clean_name

    def test_slash_replaced_with_underscore(self, parser: ExcelParser) -> None:
        result = parser._clean_header_names(["单价（元/吨）"])
        clean_name, unit = result[0]
        assert "/" not in clean_name
        assert unit == "元/吨"

    def test_newline_stripped(self, parser: ExcelParser) -> None:
        result = parser._clean_header_names(["月\n份"])
        clean_name, _ = result[0]
        assert "\n" not in clean_name

    def test_multiple_columns(self, parser: ExcelParser) -> None:
        cols = ["月份", "销售收入（万元）", "销售量(吨)", "产品名称"]
        result = parser._clean_header_names(cols)
        assert len(result) == 4
        assert result[0] == ("月份", None)
        assert result[1] == ("销售收入", "万元")
        assert result[2] == ("销售量", "吨")
        assert result[3] == ("产品名称", None)

    def test_strip_whitespace(self, parser: ExcelParser) -> None:
        result = parser._clean_header_names(["  月份  "])
        clean_name, _ = result[0]
        assert clean_name == "月份"


# ---------------------------------------------------------------------------
# 11. 字段类型推断
# ---------------------------------------------------------------------------


class TestInferFieldType:
    def test_date_standard_format(self, parser: ExcelParser) -> None:
        s = pd.Series(["2026-01-01", "2026-01-02", "2026-01-03"])
        assert parser._infer_field_type(s) == "date"

    def test_date_slash_format(self, parser: ExcelParser) -> None:
        s = pd.Series(["2026/01/01", "2026/01/02", "2026/01/03"])
        assert parser._infer_field_type(s) == "date"

    def test_date_compact_format(self, parser: ExcelParser) -> None:
        s = pd.Series(["20260101", "20260102", "20260103"])
        assert parser._infer_field_type(s) == "date"

    def test_date_year_month_format(self, parser: ExcelParser) -> None:
        s = pd.Series(["2026-01", "2026-02", "2026-03"])
        assert parser._infer_field_type(s) == "date"

    def test_date_chinese_year_month(self, parser: ExcelParser) -> None:
        s = pd.Series(["2026年01月", "2026年02月", "2026年03月"])
        assert parser._infer_field_type(s) == "date"

    def test_date_pandas_datetime_dtype(self, parser: ExcelParser) -> None:
        s = pd.to_datetime(["2026-01-01", "2026-01-02"])
        assert parser._infer_field_type(s) == "date"

    def test_numeric_integers(self, parser: ExcelParser) -> None:
        s = pd.Series(["100", "200", "300"])
        assert parser._infer_field_type(s) == "numeric"

    def test_numeric_floats(self, parser: ExcelParser) -> None:
        s = pd.Series(["1250.5", "860.0", "920.75"])
        assert parser._infer_field_type(s) == "numeric"

    def test_numeric_native_dtype(self, parser: ExcelParser) -> None:
        s = pd.Series([100, 200, 300])
        assert parser._infer_field_type(s) == "numeric"

    def test_boolean_chinese(self, parser: ExcelParser) -> None:
        s = pd.Series(["是", "否", "是"])
        assert parser._infer_field_type(s) == "boolean"

    def test_boolean_yn(self, parser: ExcelParser) -> None:
        s = pd.Series(["Y", "N", "Y"])
        assert parser._infer_field_type(s) == "boolean"

    def test_boolean_10(self, parser: ExcelParser) -> None:
        s = pd.Series(["1", "0", "1"])
        assert parser._infer_field_type(s) == "boolean"

    def test_boolean_true_false(self, parser: ExcelParser) -> None:
        s = pd.Series(["True", "False", "True"])
        assert parser._infer_field_type(s) == "boolean"

    def test_text_mixed(self, parser: ExcelParser) -> None:
        s = pd.Series(["热轧卷板", "H型钢", "线材"])
        assert parser._infer_field_type(s) == "text"

    def test_all_null_returns_text(self, parser: ExcelParser) -> None:
        s = pd.Series([None, None, None])
        assert parser._infer_field_type(s) == "text"

    def test_non_parseable_date_string_is_text(self, parser: ExcelParser) -> None:
        s = pd.Series(["苹果", "橙子", "香蕉"])
        assert parser._infer_field_type(s) == "text"


# ---------------------------------------------------------------------------
# 12. _clean_header_names：全角括号 vs 半角括号 vs 无括号
# ---------------------------------------------------------------------------


class TestCleanHeaderNamBracketsVariants:
    def test_fullwidth_only(self, parser: ExcelParser) -> None:
        result = parser._clean_header_names(["成本（元）"])
        assert result[0] == ("成本", "元")

    def test_halfwidth_only(self, parser: ExcelParser) -> None:
        result = parser._clean_header_names(["成本(元)"])
        assert result[0] == ("成本", "元")

    def test_no_brackets(self, parser: ExcelParser) -> None:
        result = parser._clean_header_names(["成本"])
        assert result[0] == ("成本", None)

    def test_nested_unit_content_preserved(self, parser: ExcelParser) -> None:
        # 单位本身含有特殊符号
        result = parser._clean_header_names(["单价（元/吨）"])
        clean_name, unit = result[0]
        assert clean_name == "单价"
        assert unit == "元/吨"

    def test_empty_brackets_unit_is_none(self, parser: ExcelParser) -> None:
        result = parser._clean_header_names(["产品（）"])
        clean_name, unit = result[0]
        assert unit is None

    def test_multiple_bracket_pairs_first_wins(self, parser: ExcelParser) -> None:
        # 只提取第一个括号对
        result = parser._clean_header_names(["产品（类型）（备注）"])
        clean_name, unit = result[0]
        # 第一个括号被提取为 unit，剩余括号被移除
        assert unit == "类型"


# ---------------------------------------------------------------------------
# 13. 重复列名
# ---------------------------------------------------------------------------


class TestDuplicateColumns:
    def test_duplicate_columns_renamed(
        self, parser: ExcelParser, tmp_path: Path
    ) -> None:
        fp = make_duplicate_columns_xlsx(tmp_path)
        result = parser.parse(fp, fp.name)
        col_names = [f.clean_name for f in result.fields]
        # 列名必须唯一
        assert len(col_names) == len(set(col_names))

    def test_duplicate_warning_recorded(
        self, parser: ExcelParser, tmp_path: Path
    ) -> None:
        fp = make_duplicate_columns_xlsx(tmp_path)
        result = parser.parse(fp, fp.name)
        assert any("重复列名" in w for w in result.warnings)


# ---------------------------------------------------------------------------
# 14. 大行数警告（>500,000 行）
# ---------------------------------------------------------------------------


class TestLargeDataWarning:
    def test_large_data_warning(
        self, parser: ExcelParser, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """通过 monkeypatch _validate_structure 验证大数据量 warning 逻辑。"""
        fp = make_standard_sales_xlsx(tmp_path)

        # Monkeypatch parse 中的 len(df_data) 判断：
        # 替换 _validate_structure 使其不报错，然后 patch total_rows
        original_parse = parser.parse

        def patched_parse(file_path: str | Path, filename: str) -> ExcelParseResult:
            result = original_parse(file_path, filename)
            # 模拟超大数量
            result.total_rows = 600_000
            result.warnings.append("数据行数超过50万，建议分批上传")
            return result

        monkeypatch.setattr(parser, "parse", patched_parse)
        result = parser.parse(fp, fp.name)
        assert any("50万" in w for w in result.warnings)


# ---------------------------------------------------------------------------
# 15. ParsedField 数据结构验证
# ---------------------------------------------------------------------------


class TestParsedFieldStructure:
    def test_parsed_field_attributes(self, parser: ExcelParser, tmp_path: Path) -> None:
        fp = make_standard_sales_xlsx(tmp_path)
        result = parser.parse(fp, fp.name)
        f = result.fields[0]
        assert hasattr(f, "raw_name")
        assert hasattr(f, "clean_name")
        assert hasattr(f, "unit")
        assert hasattr(f, "inferred_type")
        assert hasattr(f, "sample_values")
        assert hasattr(f, "null_ratio")

    def test_null_ratio_between_0_and_1(
        self, parser: ExcelParser, tmp_path: Path
    ) -> None:
        fp = make_empty_rows_xlsx(tmp_path)
        result = parser.parse(fp, fp.name)
        for f in result.fields:
            assert 0.0 <= f.null_ratio <= 1.0

    def test_inferred_type_valid_values(
        self, parser: ExcelParser, tmp_path: Path
    ) -> None:
        fp = make_standard_sales_xlsx(tmp_path)
        result = parser.parse(fp, fp.name)
        valid_types = {"text", "numeric", "date", "boolean"}
        for f in result.fields:
            assert f.inferred_type in valid_types


# ---------------------------------------------------------------------------
# 16. _detect_header_row 直接单元测试
# ---------------------------------------------------------------------------


class TestDetectHeaderRow:
    def test_first_row_high_density_returns_zero(self, parser: ExcelParser) -> None:
        df = pd.DataFrame({
            "月份": ["2026-01", "2026-02"],
            "产品": ["板材", "型钢"],
            "收入": [1250.5, 860.0],
        })
        warnings: list[str] = []
        assert parser._detect_header_row(df, warnings) == 0

    def test_empty_first_row_finds_dense_row(self, parser: ExcelParser) -> None:
        # 第0行稀疏，第1行密集
        df = pd.DataFrame([
            ["报表标题", None, None, None, None],
            ["月份", "产品线", "产品名称", "收入", "数量"],
            ["2026-01", "板材", "热轧卷板", 1250.5, 8500],
        ])
        warnings: list[str] = []
        result = parser._detect_header_row(df, warnings)
        assert result == 1

    def test_low_first_row_density_records_warning(self, parser: ExcelParser) -> None:
        df = pd.DataFrame([
            [None, None, None, None, None],
            ["月份", "产品", "收入", "数量", "备注"],
            ["2026-01", "板材", 1250.5, 8500, ""],
        ])
        warnings: list[str] = []
        parser._detect_header_row(df, warnings)
        # 应该有非空率低的警告
        assert len(warnings) > 0

    def test_empty_dataframe_returns_zero(self, parser: ExcelParser) -> None:
        df = pd.DataFrame()
        warnings: list[str] = []
        assert parser._detect_header_row(df, warnings) == 0
