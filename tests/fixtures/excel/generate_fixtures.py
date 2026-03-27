"""
生成单元测试所需的 Excel fixtures

使用方法：
    cd backend
    python -m tests.fixtures.excel.generate_fixtures
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pandas as pd

OUTPUT_DIR = Path(__file__).parent


def generate_standard_sales() -> None:
    """标准格式销售台账"""
    data = {
        "月份": ["2026-01", "2026-01", "2026-02", "2026-02", "2026-03", "2026-03"],
        "产品线": ["板材", "型钢", "板材", "型钢", "板材", "型钢"],
        "产品名称": ["热轧卷板", "H型钢", "热轧卷板", "H型钢", "热轧卷板", "H型钢"],
        "销售收入（万元）": [1250.50, 860.00, 1380.20, 920.00, 1320.00, 900.00],
        "销售量（吨）": [8500, 6200, 9200, 6800, 8900, 6500],
        "单价（元/吨）": [1471.18, 1387.10, 1500.22, 1352.94, 1483.15, 1384.62],
        "客户名称": ["客户A", "客户B", "客户A", "客户C", "客户B", "客户A"],
    }
    df = pd.DataFrame(data)
    df.to_excel(OUTPUT_DIR / "standard_sales.xlsx", index=False)
    print("✅ 生成 standard_sales.xlsx")


def generate_merged_cells() -> None:
    """含合并单元格的表格"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "销售数据"

    # 表头行（含合并单元格）
    ws["A1"] = "报告月份"
    ws["B1"] = "产品信息"
    ws.merge_cells("B1:C1")  # 合并 B1:C1
    ws["D1"] = "金额（万元）"
    ws["E1"] = "数量（吨）"

    ws["A2"] = "月份"
    ws["B2"] = "产品线"
    ws["C2"] = "产品名称"
    ws["D2"] = "收入"
    ws["E2"] = "销量"

    # 数据行（月份列合并）
    ws["A3"] = "2026-01"
    ws.merge_cells("A3:A4")
    ws["B3"] = "板材"
    ws["C3"] = "热轧卷板"
    ws["D3"] = 1250.50
    ws["E3"] = 8500

    ws["B4"] = "型钢"
    ws["C4"] = "H型钢"
    ws["D4"] = 860.00
    ws["E4"] = 6200

    wb.save(OUTPUT_DIR / "merged_cells.xlsx")
    print("✅ 生成 merged_cells.xlsx")


def generate_mixed_types() -> None:
    """混合类型字段"""
    data = {
        "日期": ["2026-01-01", "2026年1月", "2026/01/15", "20260120", "2026-02"],
        "金额": [1250.5, "1380万", "920元", "860.0", 1000],
        "产品": ["热轧卷板", "H型钢", None, "线材", "螺纹钢"],
        "是否完成": ["是", "否", "Y", "N", "1"],
        "数量": ["8500吨", 6200, "9200", "未知", 8900],
    }
    df = pd.DataFrame(data)
    df.to_excel(OUTPUT_DIR / "mixed_types.xlsx", index=False)
    print("✅ 生成 mixed_types.xlsx")


def generate_empty_rows() -> None:
    """含空行和空列"""
    data = {
        "月份": ["2026-01", None, "2026-02", None, "2026-03"],
        "产品线": ["板材", None, "型钢", None, "板材"],
        "收入（万元）": [1250.5, None, 860.0, None, 1380.2],
        "备注": [None, None, None, None, None],  # 全空列
    }
    df = pd.DataFrame(data)
    df.to_excel(OUTPUT_DIR / "empty_rows.xlsx", index=False)
    print("✅ 生成 empty_rows.xlsx")


def generate_bad_encoding() -> None:
    """GBK 编码 CSV"""
    content = "月份,产品线,收入（万元）\n2026-01,板材,1250.5\n2026-01,型钢,860.0\n"
    (OUTPUT_DIR / "bad_encoding.csv").write_bytes(content.encode("gbk"))
    print("✅ 生成 bad_encoding.csv")


def generate_multi_header() -> None:
    """多级表头（非标准格式）"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "财务数据"

    # 第一行：标题行（应被跳过）
    ws["A1"] = "2026年1-3月财务汇总报表"
    ws.merge_cells("A1:E1")

    # 第二行：主表头
    ws["A2"] = "月份"
    ws["B2"] = "产品名称"
    ws["C2"] = "收入"
    ws["D2"] = "成本"
    ws["E2"] = "毛利"

    # 数据从第三行开始
    ws["A3"] = "2026-01"
    ws["B3"] = "热轧卷板"
    ws["C3"] = 1250.5
    ws["D3"] = 1050.0
    ws["E3"] = 200.5

    ws["A4"] = "2026-02"
    ws["B4"] = "热轧卷板"
    ws["C4"] = 1380.2
    ws["D4"] = 1150.0
    ws["E4"] = 230.2

    wb.save(OUTPUT_DIR / "multi_header.xlsx")
    print("✅ 生成 multi_header.xlsx")


if __name__ == "__main__":
    generate_standard_sales()
    generate_merged_cells()
    generate_mixed_types()
    generate_empty_rows()
    generate_bad_encoding()
    generate_multi_header()
    print("\n所有 Excel fixtures 生成完毕。")
