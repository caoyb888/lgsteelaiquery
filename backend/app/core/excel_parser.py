"""
Excel 解析引擎

支持 .xlsx / .xls / .csv 格式，自动检测表头行，展开合并单元格，
推断字段类型，校验结构完整性，返回 ExcelParseResult 供后续字段映射使用。

在 Celery 任务中同步调用（非 async），IO 密集但在 worker 进程中执行。
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl
import pandas as pd
from loguru import logger

from app.config import get_settings
from app.utils.exceptions import (
    ExcelFileTooLargeError,
    ExcelParseError,
    UnsupportedFormatError,
)

# ---------------------------------------------------------------------------
# 数据类
# ---------------------------------------------------------------------------


@dataclass
class ParsedField:
    raw_name: str           # 原始列名
    clean_name: str         # 清理后列名（去除单位后缀括号，去空格）
    unit: str | None        # 提取的单位，如"万元"、"吨"
    inferred_type: str      # "text" | "numeric" | "date" | "boolean"
    sample_values: list     # 前3个非空值（供字段映射参考）
    null_ratio: float       # 空值率


@dataclass
class ExcelParseResult:
    df: pd.DataFrame                    # 清洗后的 DataFrame（仅包含有效列）
    fields: list[ParsedField]           # 字段元信息列表
    header_row_index: int               # 检测到的表头行（0-based）
    sheet_name: str                     # 工作表名
    total_rows: int                     # 数据行数（不含表头）
    warnings: list[str]                 # 结构警告列表
    source_filename: str                # 原始文件名


# ---------------------------------------------------------------------------
# 主类
# ---------------------------------------------------------------------------

# 匹配全角或半角括号内的单位，如（万元）、(吨)
_UNIT_RE = re.compile(r"[（(]([^）)]+)[）)]")

# 用于 SQL 列名安全化的字符替换
_UNSAFE_CHARS_RE = re.compile(r"[\s/\\.]")


class ExcelParser:
    """
    Excel / CSV 文件解析器。

    同步解析入口 `parse()` 被 Celery 任务调用，执行文件校验、格式路由、
    表头检测、合并单元格展开、字段类型推断和结构校验，返回 ExcelParseResult。
    """

    # 尝试推断日期的格式列表（按优先级）
    _DATE_FORMATS: list[str] = [
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%Y%m%d",
        "%Y-%m",
        "%Y年%m月%d日",
        "%Y年%m月",
    ]

    # 视为布尔值的字符串集合（大小写不敏感，全部转小写后匹配）
    _BOOL_VALUES: frozenset[str] = frozenset(
        {"是", "否", "y", "n", "1", "0", "true", "false"}
    )

    def __init__(self) -> None:
        self._settings = get_settings()

    # ------------------------------------------------------------------
    # 公开接口
    # ------------------------------------------------------------------

    def parse(self, file_path: str | Path, filename: str) -> ExcelParseResult:
        """
        同步解析入口（在 Celery 任务中调用）。

        步骤：
        1. 文件大小校验
        2. 文件格式校验
        3. 根据后缀路由到 _parse_xlsx / _parse_xls / _parse_csv
        4. 检测表头行
        5. 展开合并单元格（仅 xlsx）
        6. 清理表头名称
        7. 推断字段类型
        8. 校验结构
        9. 组装并返回 ExcelParseResult
        """
        fp = Path(file_path)
        suffix = fp.suffix.lower()

        logger.info("开始解析 Excel 文件", filename=filename, suffix=suffix)

        # 1. 文件大小校验
        file_size = fp.stat().st_size
        max_bytes = self._settings.excel_max_size_bytes
        if file_size > max_bytes:
            raise ExcelFileTooLargeError(
                f"文件 {filename} 大小 {file_size} 字节，超过限制 {max_bytes} 字节"
            )

        # 2. 文件格式校验
        supported = {".xlsx", ".xls", ".csv"}
        if suffix not in supported:
            raise UnsupportedFormatError(
                f"不支持的文件格式 {suffix!r}，请上传 .xlsx / .xls / .csv 文件"
            )

        warnings: list[str] = []

        # 3. 路由到具体解析方法（先得到原始 DataFrame）
        if suffix == ".xlsx":
            df_raw, sheet_name = self._read_xlsx_raw(fp)
        elif suffix == ".xls":
            df_raw, sheet_name = self._read_xls_raw(fp)
        else:
            df_raw, sheet_name = self._read_csv_raw(fp)

        # 4. 检测表头行
        header_row_index = self._detect_header_row(df_raw, warnings)

        # 重新设置表头：取 header_row_index 行作为列名，数据从下一行开始
        new_columns = df_raw.iloc[header_row_index].tolist()
        df_data = df_raw.iloc[header_row_index + 1 :].reset_index(drop=True)
        df_data.columns = pd.Index(
            [str(c) if c is not None else f"unnamed_{i}" for i, c in enumerate(new_columns)]
        )

        # 5. 展开合并单元格（仅 xlsx，已在 _read_xlsx_raw 中完成）
        #    对于 xlsx，我们已在读取阶段展开；此处无需额外操作。

        # 6. 清理表头名称
        raw_columns = df_data.columns.tolist()
        cleaned = self._clean_header_names([str(c) for c in raw_columns])

        # 重命名 DataFrame 列为清理后的名称
        rename_map = {
            old: clean_name
            for old, (clean_name, _) in zip(raw_columns, cleaned)
        }
        df_data = df_data.rename(columns=rename_map)

        # 7. 校验结构（全空列、重复列名，数据行数）
        # 保存 validate 前的列顺序（作为 raw_name 映射依据）
        pre_validate_cols = list(df_data.columns)
        self._validate_structure(df_data, warnings)

        # 8. 推断字段类型，组装 ParsedField 列表
        # 以 validate_structure 后的 df_data.columns 为准（去重后的列名可能已变化）
        # 建立 pre_validate 列名 → (raw_name, unit) 的反向映射
        clean_to_raw: dict[str, tuple[str, str | None]] = {}
        for raw_name, (clean_name, unit) in zip(raw_columns, cleaned):
            if clean_name not in clean_to_raw:
                clean_to_raw[clean_name] = (raw_name, unit)

        fields: list[ParsedField] = []
        for actual_col in df_data.columns:
            # 去掉自动加的数字后缀（_1, _2 ...）找回原始 raw_name/unit
            base_col = actual_col.rsplit("_", 1)[0] if "_" in actual_col else actual_col
            raw_name, unit = clean_to_raw.get(actual_col) or clean_to_raw.get(base_col) or (actual_col, None)
            series = df_data[actual_col]
            inferred_type = self._infer_field_type(series)
            non_null = series.dropna()
            sample_values = non_null.head(3).tolist()
            null_ratio = series.isna().mean()
            fields.append(
                ParsedField(
                    raw_name=raw_name,
                    clean_name=actual_col,
                    unit=unit,
                    inferred_type=inferred_type,
                    sample_values=sample_values,
                    null_ratio=float(null_ratio),
                )
            )

        total_rows = len(df_data)

        if total_rows > 500_000:
            warnings.append("数据行数超过50万，建议分批上传")

        logger.info(
            "Excel 解析完成",
            filename=filename,
            total_rows=total_rows,
            fields=len(fields),
            warnings=len(warnings),
        )

        return ExcelParseResult(
            df=df_data,
            fields=fields,
            header_row_index=header_row_index,
            sheet_name=sheet_name,
            total_rows=total_rows,
            warnings=warnings,
            source_filename=filename,
        )

    # ------------------------------------------------------------------
    # 格式读取（原始 DataFrame，不处理表头）
    # ------------------------------------------------------------------

    def _read_xlsx_raw(self, fp: Path) -> tuple[pd.DataFrame, str]:
        """
        读取 xlsx 文件，先用 openpyxl 展开合并单元格，再转为 DataFrame。
        """
        try:
            wb = openpyxl.load_workbook(fp, data_only=True)
            ws = wb.active
            sheet_name: str = ws.title if ws.title else "Sheet1"  # type: ignore[assignment]
            df = self._expand_merged_cells(wb, sheet_name)
            wb.close()
        except Exception as exc:
            raise ExcelParseError(f"xlsx 文件读取失败：{exc}") from exc
        return df, sheet_name

    def _read_xls_raw(self, fp: Path) -> tuple[pd.DataFrame, str]:
        """读取旧版 xls 文件（xlrd 引擎，不处理合并单元格）。"""
        try:
            xl = pd.ExcelFile(fp, engine="xlrd")
            sheet_name = xl.sheet_names[0]
            df = xl.parse(sheet_name, header=None, dtype=str, keep_default_na=False)
            df = df.replace("", None)
        except Exception as exc:
            raise ExcelParseError(f"xls 文件读取失败：{exc}") from exc
        return df, sheet_name

    def _read_csv_raw(self, fp: Path) -> tuple[pd.DataFrame, str]:
        """
        读取 CSV 文件，自动检测编码（优先 UTF-8，fallback GBK）。
        """
        for encoding in ("utf-8-sig", "utf-8", "gbk", "gb2312"):
            try:
                df = pd.read_csv(
                    fp,
                    header=None,
                    dtype=str,
                    keep_default_na=False,
                    encoding=encoding,
                )
                df = df.replace("", None)
                logger.debug("CSV 编码检测成功", encoding=encoding, filename=fp.name)
                return df, "Sheet1"
            except (UnicodeDecodeError, UnicodeError):
                continue
            except Exception as exc:
                raise ExcelParseError(f"CSV 文件读取失败：{exc}") from exc
        raise ExcelParseError(
            f"CSV 文件编码无法识别，请转换为 UTF-8 或 GBK 后重新上传：{fp.name}"
        )

    # ------------------------------------------------------------------
    # 合并单元格展开
    # ------------------------------------------------------------------

    def _expand_merged_cells(self, wb: openpyxl.Workbook, sheet_name: str) -> pd.DataFrame:
        """
        展开工作表中的所有合并单元格区域：
        将合并区域左上角单元格的值填充到区域内所有格，
        然后将工作表数据转换为 DataFrame（全部列为 str 或 None）。
        """
        ws = wb[sheet_name]

        # 收集需要填充的合并区域及对应值
        # openpyxl 中合并格的非左上角单元格读取会报 MergedCell 类型，不含 value
        merged_ranges = list(ws.merged_cells.ranges)
        fill_map: dict[tuple[int, int], object] = {}

        for merged_range in merged_ranges:
            # 左上角坐标（1-based）
            min_row = merged_range.min_row
            min_col = merged_range.min_col
            fill_value = ws.cell(row=min_row, column=min_col).value
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    if (row, col) != (min_row, min_col):
                        fill_map[(row, col)] = fill_value

        # 取消合并（unmerge）后才能安全写入
        for merged_range in merged_ranges:
            ws.unmerge_cells(str(merged_range))

        # 将填充值写回
        for (row, col), value in fill_map.items():
            ws.cell(row=row, column=col, value=value)

        # 转换为 DataFrame
        data = list(ws.values)
        if not data:
            return pd.DataFrame()

        # 全部转换为字符串（保持与 read_excel header=None 一致的行为）
        rows = []
        for row in data:
            rows.append([str(v) if v is not None else None for v in row])

        max_cols = max(len(r) for r in rows) if rows else 0
        # 补齐短行
        for r in rows:
            while len(r) < max_cols:
                r.append(None)

        df = pd.DataFrame(rows)
        return df

    # ------------------------------------------------------------------
    # 表头行检测
    # ------------------------------------------------------------------

    def _detect_header_row(self, df_raw: pd.DataFrame, warnings: list[str]) -> int:
        """
        扫描前 10 行，找非空率最高的行作为表头行（0-based index）。

        快速路径：若第0行非空率 ≥ 80%，直接返回 0。
        """
        scan_rows = min(10, len(df_raw))
        if scan_rows == 0:
            return 0

        def non_null_ratio(row_idx: int) -> float:
            row = df_raw.iloc[row_idx]
            total = len(row)
            if total == 0:
                return 0.0
            non_null = row.notna().sum() + (row == "").sum()  # type: ignore[operator]
            # 真正非空：不是 None 且不是空字符串
            truly_non_null = sum(
                1 for v in row if v is not None and str(v).strip() != ""
            )
            return truly_non_null / total

        first_row_ratio = non_null_ratio(0)

        if first_row_ratio < 0.3:
            warnings.append(
                f"第0行非空率仅 {first_row_ratio:.0%}，可能存在标题行，已尝试自动检测表头"
            )

        if first_row_ratio >= 0.8:
            return 0

        # 扫描前 scan_rows 行，找最高非空率的行
        best_row = 0
        best_ratio = first_row_ratio
        for i in range(1, scan_rows):
            ratio = non_null_ratio(i)
            if ratio > best_ratio:
                best_ratio = ratio
                best_row = i

        return best_row

    # ------------------------------------------------------------------
    # 表头名清理
    # ------------------------------------------------------------------

    def _clean_header_names(
        self, columns: list[str]
    ) -> list[tuple[str, str | None]]:
        """
        清理列名列表，返回 [(clean_name, unit_or_None), ...]。

        规则：
        - 去首尾空格
        - 去除换行符
        - 提取括号（全角/半角）内的单位
        - 将列名中的空格、/、\\、. 替换为 _（用于 SQL 列名）
        """
        result: list[tuple[str, str | None]] = []
        for col in columns:
            name = col.replace("\n", "").replace("\r", "").strip()

            # 提取单位
            unit: str | None = None
            match = _UNIT_RE.search(name)
            if match:
                unit = match.group(1).strip() or None
                # 去掉括号及其内容
                name = _UNIT_RE.sub("", name).strip()

            # 替换 SQL 不安全字符
            clean_name = _UNSAFE_CHARS_RE.sub("_", name).strip("_")
            # 去掉多余连续下划线
            clean_name = re.sub(r"_+", "_", clean_name)

            result.append((clean_name, unit))
        return result

    # ------------------------------------------------------------------
    # 字段类型推断
    # ------------------------------------------------------------------

    def _infer_field_type(self, series: pd.Series) -> str:
        """
        对单列推断数据类型，返回 "date" / "numeric" / "boolean" / "text"。

        取非空样本（最多 20 个），按优先级尝试：
        1. pandas dtype 是 datetime → "date"
        2. 尝试多种中文日期格式 → "date"
        3. 尝试 pd.to_numeric → "numeric"
        4. 样本均在布尔值集合中 → "boolean"
        5. 否则 → "text"
        """
        # 1. dtype 直接是 datetime
        if pd.api.types.is_datetime64_any_dtype(series):
            return "date"

        sample = series.dropna().head(20)
        if sample.empty:
            return "text"

        str_sample = sample.astype(str).tolist()

        # 2. 尝试日期格式
        if self._try_parse_dates(str_sample):
            return "date"

        # 3. 布尔值（优先于数值，避免"1"/"0"被误判为 numeric）
        if all(v.strip().lower() in self._BOOL_VALUES for v in str_sample):
            return "boolean"

        # 4. 尝试数值
        try:
            pd.to_numeric(pd.Series(str_sample))
            return "numeric"
        except (ValueError, TypeError):
            pass

        return "text"

    def _try_parse_dates(self, str_values: list[str]) -> bool:
        """
        对字符串样本列表尝试多种日期格式，若全部可解析则返回 True。
        至少需要 1 个样本值。
        """
        if not str_values:
            return False

        for fmt in self._DATE_FORMATS:
            parsed = 0
            for v in str_values:
                try:
                    pd.to_datetime(v, format=fmt)
                    parsed += 1
                except (ValueError, TypeError):
                    break
            if parsed == len(str_values):
                return True
        return False

    # ------------------------------------------------------------------
    # 结构校验
    # ------------------------------------------------------------------

    def _validate_structure(self, df: pd.DataFrame, warnings: list[str]) -> None:
        """
        原地修改 df，校验并处理：
        - 全空列：记录 warning，drop 该列
        - 重复列名：记录 warning，自动重命名
        - 数据行数 < 1：抛 ExcelParseError
        - 数据行数 > 500000：记录 warning（在 parse() 中处理）
        """
        # --- 全空列 ---
        cols_to_drop: list[str] = []
        for i, col in enumerate(df.columns):
            # 用位置索引避免重复列名引起的 Series 歧义
            series = df.iloc[:, i]
            if bool(series.isna().all()):
                warnings.append(f"列 {col!r} 全部为空，已跳过")
                cols_to_drop.append(col)
        if cols_to_drop:
            df.drop(columns=cols_to_drop, inplace=True)

        # --- 重复列名 ---
        seen: dict[str, int] = {}
        new_columns: list[str] = []
        has_dup = False
        for col in df.columns:
            col_str = str(col)
            if col_str in seen:
                seen[col_str] += 1
                new_col = f"{col_str}_{seen[col_str]}"
                new_columns.append(new_col)
                has_dup = True
            else:
                seen[col_str] = 0
                new_columns.append(col_str)
        if has_dup:
            dup_names = {c for c, cnt in seen.items() if cnt > 0}
            for name in dup_names:
                warnings.append(f"发现重复列名 {name!r}，已自动重命名")
            df.columns = pd.Index(new_columns)

        # --- 数据行数 ---
        if len(df) < 1:
            raise ExcelParseError("Excel 文件无有效数据行")
